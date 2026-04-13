import os
from datetime import datetime, timedelta
from pathlib import Path
from dotenv import load_dotenv

load_dotenv()

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ============================================================
# CONFIGURATION
#
# Set your Polygon key before running:
#   Windows PowerShell: $env:POLYGON_API_KEY="your_key_here"
#   macOS/Linux:        export POLYGON_API_KEY="your_key_here"
#
# Set DM_CSV_PATH to wherever Avinash's DM_Latest CSV lives.
# Set DM_HISTORY_PATH to the accumulated daily DM history CSV.
#   Expected columns: Date (YYYY-MM-DD), Ticker, DM, Phase, Close
#   One row per ticker per session -- same structure as dm_latest
#   exported daily and appended.
# ============================================================

API_KEY     = os.getenv("POLYGON_API_KEY", "")
API_URL     = "https://api.polygon.io/v2/snapshot/locale/us/markets/stocks/gainers"
OUTPUT_DIR  = Path("market_snapshots_polygon")
TOP_N       = 20

# Lookback window in trading days
LOOKBACK_DAYS    = int(os.getenv("DM_LOOKBACK_DAYS", "90"))

# DM classification thresholds
DM_STRONG   = 65   # Institutionally backed -- enter candidate
DM_NEUTRAL  = 40   # Mixed -- watch only
# Below DM_NEUTRAL = head fake territory


# ============================================================
# DM LOADER -- current snapshot (from Supabase dm_latest)
# ============================================================

def _get_supabase():
    url = os.getenv("SUPABASE_URL")
    key = os.getenv("SUPABASE_KEY")
    if not url or not key:
        return None
    from supabase import create_client
    return create_client(url, key)


def load_dm_data() -> dict:
    """Load DM data from Supabase dm_latest table."""
    client = _get_supabase()
    if not client:
        print("[WARN] SUPABASE_URL/KEY not set -- running without DM data.")
        return {}

    try:
        rows = []
        offset, page = 0, 5000
        while True:
            r = (client.table('dm_latest')
                 .select('ticker,dm_smoothed,dm_change,phase,close')
                 .range(offset, offset + page - 1)
                 .execute())
            rows.extend(r.data)
            if len(r.data) < page:
                break
            offset += page

        result = {}
        for row in rows:
            ticker = str(row['ticker']).strip().upper()
            result[ticker] = {
                "dm":       float(row['dm_smoothed']) if row.get('dm_smoothed') is not None else None,
                "dm_chg":   float(row['dm_change'])   if row.get('dm_change') is not None else None,
                "phase":    str(row.get('phase', '')),
                "dm_close": float(row['close'])       if row.get('close') is not None else None,
            }
        print(f"[DM] Loaded {len(result)} tickers from Supabase dm_latest")
        return result
    except Exception as e:
        print(f"[WARN] Failed to load DM data: {e}")
        return {}


# ============================================================
# DM HISTORY LOADER -- lookback (from Supabase dm_history)
# ============================================================

def load_dm_history() -> pd.DataFrame:
    """Load DM history from Supabase dm_history table for the lookback window."""
    client = _get_supabase()
    if not client:
        print("[WARN] SUPABASE not available -- lookback disabled.")
        return None

    try:
        cutoff = (datetime.now() - timedelta(days=int(LOOKBACK_DAYS * 1.5))).strftime('%Y-%m-%d')
        print(f"[DM History] Loading from Supabase (since {cutoff})...")

        rows = []
        offset, page = 0, 10000
        while True:
            r = (client.table('dm_history')
                 .select('date,ticker,dm_smoothed,phase,close')
                 .gte('date', cutoff)
                 .order('date')
                 .range(offset, offset + page - 1)
                 .execute())
            rows.extend(r.data)
            if len(r.data) < page:
                break
            offset += page

        if not rows:
            print("[WARN] No DM history returned -- lookback disabled.")
            return None

        df = pd.DataFrame(rows)
        df.rename(columns={'dm_smoothed': 'DM', 'ticker': 'Ticker', 'date': 'Date', 'close': 'Close', 'phase': 'Phase'}, inplace=True)
        df['Date'] = pd.to_datetime(df['Date'])
        df['DM'] = pd.to_numeric(df['DM'], errors='coerce')
        df['Close'] = pd.to_numeric(df['Close'], errors='coerce')
        df['Ticker'] = df['Ticker'].str.strip().str.upper()
        df = df.sort_values(['Ticker', 'Date'])

        print(f"[DM History] Loaded {len(df):,} rows, "
              f"{df['Ticker'].nunique()} tickers, "
              f"{df['Date'].min().date()} -> {df['Date'].max().date()}")
        return df
    except Exception as e:
        print(f"[WARN] Failed to load DM history: {e}")
        return None


def find_entry_signal(history_df: pd.DataFrame, ticker: str,
                      current_price: float, lookback_days: int) -> dict:
    """
    For a given ticker, scan the lookback window to find the FIRST date
    DM crossed DM_STRONG (65). Returns entry analysis dict.

    Returns:
        entry_date       -- date DM first crossed 65 (or None)
        entry_dm         -- DM value on entry date
        entry_price      -- Close price on entry date (from history)
        days_held        -- calendar days from entry to today
        simulated_return -- % return if entered at entry_close, exited at current_price
        signal_type      -- STRONG_ENTRY / NEVER_CROSSED / NOT_IN_HISTORY / NO_HISTORY
        peak_dm          -- highest DM in the lookback window
        days_above_gate  -- sessions where DM >= 65 in window
    """
    empty = {
        "entry_date": None, "entry_dm": None, "entry_price": None,
        "days_held": None, "simulated_return": None, "signal_type": "NO_HISTORY",
        "peak_dm": None, "days_above_gate": None,
    }

    if history_df is None:
        return empty

    tk_df = history_df[history_df["Ticker"] == ticker].copy()
    if tk_df.empty:
        empty["signal_type"] = "NOT_IN_HISTORY"
        return empty

    # Restrict to lookback window (use 1.5x calendar days for trading day approximation)
    cutoff = datetime.now() - timedelta(days=lookback_days * 1.5)
    tk_df = tk_df[tk_df["Date"] >= cutoff].reset_index(drop=True)
    if tk_df.empty:
        empty["signal_type"] = "NO_HISTORY"
        return empty

    # Find first date DM >= DM_STRONG
    above        = tk_df[tk_df["DM"] >= DM_STRONG]
    peak_dm      = float(tk_df["DM"].max())
    days_above   = int((tk_df["DM"] >= DM_STRONG).sum())

    if above.empty:
        return {**empty,
                "signal_type":     "NEVER_CROSSED",
                "peak_dm":         peak_dm,
                "days_above_gate": 0}

    entry_row   = above.iloc[0]
    entry_date  = entry_row["Date"]
    entry_dm    = float(entry_row["DM"])
    entry_price = (float(entry_row["Close"])
                   if "Close" in entry_row and pd.notna(entry_row["Close"])
                   else None)

    days_held  = (datetime.now() - entry_date).days

    sim_return = None
    if entry_price and entry_price > 0 and current_price:
        sim_return = (current_price - entry_price) / entry_price * 100

    return {
        "entry_date":        entry_date.strftime("%Y-%m-%d"),
        "entry_dm":          round(entry_dm, 1),
        "entry_price":       round(entry_price, 2) if entry_price else None,
        "days_held":         days_held,
        "simulated_return":  round(sim_return, 1) if sim_return is not None else None,
        "signal_type":       "STRONG_ENTRY",
        "peak_dm":           round(peak_dm, 1),
        "days_above_gate":   days_above,
    }


def classify_dm(dm_score) -> str:
    if dm_score is None:
        return "NO DATA"
    if dm_score >= DM_STRONG:
        return "BACKED"       # Institutional flow confirmed -- real move
    if dm_score >= DM_NEUTRAL:
        return "WATCH"        # Mixed -- monitor, don't chase
    return "HEAD FAKE"        # Price up, institutions leaving -- avoid


# ============================================================
# POLYGON FETCH
# ============================================================

def _pick_rows(payload):
    if isinstance(payload, list):
        return payload
    if isinstance(payload, dict):
        for key in ("tickers", "results"):
            if isinstance(payload.get(key), list):
                return payload[key]
    raise RuntimeError(f"Unexpected Polygon response shape: {payload}")


def fetch_top_gainers(limit: int = TOP_N) -> pd.DataFrame:
    if not API_KEY:
        raise RuntimeError(
            "Missing POLYGON_API_KEY environment variable. "
            "Set it first, then rerun the script."
        )

    resp = requests.get(API_URL, params={"apiKey": API_KEY}, timeout=30)
    resp.raise_for_status()
    rows = _pick_rows(resp.json())
    if not rows:
        raise RuntimeError("Polygon returned no gainers rows.")

    dm_data    = load_dm_data()
    history_df = load_dm_history()

    normalized = []
    run_ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for idx, item in enumerate(rows[:limit], start=1):
        day        = item.get("day", {}) or {}
        prev_day   = item.get("prevDay", {}) or {}
        last_trade = item.get("lastTrade", {}) or {}
        min_bar    = item.get("min", {}) or {}

        symbol     = item.get("ticker") or item.get("symbol")
        change_pct = item.get("todaysChangePerc")
        if change_pct is not None:
            change_pct = float(change_pct) / 100.0

        current_price = day.get("c")

        # Current DM enrichment
        dm_row    = dm_data.get(symbol.upper(), {}) if symbol else {}
        dm_score  = dm_row.get("dm")
        dm_change = dm_row.get("dm_chg")
        dm_phase  = dm_row.get("phase", "")
        signal    = classify_dm(dm_score)

        # Lookback analysis
        entry = (find_entry_signal(history_df, symbol.upper(),
                                   current_price, LOOKBACK_DAYS)
                 if symbol else {
                     "entry_date": None, "entry_dm": None, "entry_price": None,
                     "days_held": None, "simulated_return": None,
                     "signal_type": "NO_HISTORY", "peak_dm": None,
                     "days_above_gate": None
                 })

        normalized.append({
            "Rank":              idx,
            "Symbol":            symbol,
            "Signal":            signal,
            "DM Score":          dm_score,
            "DM Change":         dm_change,
            "DM Phase":          dm_phase,
            "Price":             current_price,
            "Today Change %":    change_pct,
            "Today Change $":    item.get("todaysChange"),
            "Day Open":          day.get("o"),
            "Day High":          day.get("h"),
            "Day Low":           day.get("l"),
            "Day Volume":        day.get("v"),
            "Prev Close":        prev_day.get("c"),
            "Last Trade":        last_trade.get("p"),
            "Minute VWAP":       min_bar.get("vw"),
            # Lookback columns
            "LB Signal":         entry["signal_type"],
            "LB Entry Date":     entry["entry_date"],
            "LB Entry DM":       entry["entry_dm"],
            "LB Entry Price":    entry["entry_price"],
            "LB Days Held":      entry["days_held"],
            "LB Sim Return %":   entry["simulated_return"],
            "LB Peak DM":        entry["peak_dm"],
            "LB Days Above 65":  entry["days_above_gate"],
            "Retrieved At":      run_ts,
        })

    return pd.DataFrame(normalized)


# ============================================================
# CONSOLE SUMMARY
# ============================================================

def print_summary(df: pd.DataFrame):
    print()
    print("=" * 75)
    print(f"  TOP GAINERS -- DM CROSS-REFERENCE + {LOOKBACK_DAYS}D LOOKBACK  "
          f"{datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print(f"  Entry gate: DM >= {DM_STRONG} | Head fake: DM < {DM_NEUTRAL}")
    print("=" * 75)

    for signal_label, group_label in [
        ("BACKED",    f"INSTITUTIONALLY BACKED (DM >= {DM_STRONG})"),
        ("WATCH",     f"WATCH -- MIXED SIGNAL (DM {DM_NEUTRAL}-{DM_STRONG})"),
        ("HEAD FAKE", f"HEAD FAKES -- DO NOT CHASE (DM < {DM_NEUTRAL})"),
    ]:
        grp = df[df["Signal"] == signal_label]
        if grp.empty:
            continue
        icon = {"BACKED": "OK", "WATCH": "WATCH", "HEAD FAKE": "FAKE"}[signal_label]
        print(f"\n[{icon}]  {group_label} -- {len(grp)} names:")
        hdr  = f"    {'Sym':<7} {'Chg%':>7}  {'DM':>6}  {'DM Chg':>8}  "
        hdr += f"{'LB Signal':<16} {'Entry Dt':<12} {'Entry$':>8}  "
        hdr += f"{'SimRet%':>8}  {'DaysHeld':>9}  {'PeakDM':>7}"
        print(hdr)
        print("    " + "-" * (len(hdr) - 4))
        for _, r in grp.iterrows():
            chg    = f"{r['Today Change %']:.1%}"    if pd.notna(r['Today Change %'])   else "N/A"
            dm     = f"{r['DM Score']:.1f}"           if pd.notna(r['DM Score'])          else "--"
            dmc    = f"{r['DM Change']:+.1f}"         if pd.notna(r['DM Change'])         else "--"
            lb_sig = str(r['LB Signal'])              if pd.notna(r['LB Signal'])         else "--"
            edate  = str(r['LB Entry Date'])          if pd.notna(r['LB Entry Date'])     else "--"
            eprice = f"${r['LB Entry Price']:.2f}"    if pd.notna(r['LB Entry Price'])    else "--"
            simret = f"{r['LB Sim Return %']:+.1f}%"  if pd.notna(r['LB Sim Return %'])   else "--"
            dheld  = str(int(r['LB Days Held']))      if pd.notna(r['LB Days Held'])      else "--"
            peakdm = f"{r['LB Peak DM']:.1f}"         if pd.notna(r['LB Peak DM'])        else "--"
            print(f"    {r['Symbol']:<7} {chg:>7}  {dm:>6}  {dmc:>8}  "
                  f"{lb_sig:<16} {edate:<12} {eprice:>8}  "
                  f"{simret:>8}  {dheld:>9}  {peakdm:>7}")

    no_data = df[df["Signal"] == "NO DATA"]
    if not no_data.empty:
        print(f"\n[--]  NOT IN DM UNIVERSE ({len(no_data)}): "
              f"{', '.join(no_data['Symbol'].tolist())}")

    # Summary stats
    backed         = df[df["Signal"] == "BACKED"]
    watch          = df[df["Signal"] == "WATCH"]
    hf             = df[df["Signal"] == "HEAD FAKE"]
    strong_entries = df[df["LB Signal"] == "STRONG_ENTRY"]
    never_crossed  = df[df["LB Signal"] == "NEVER_CROSSED"]

    print()
    print(f"  CURRENT SIGNAL: {len(backed)} backed | "
          f"{len(watch)} watch | {len(hf)} head fakes")

    lb_with_data = df[df["LB Sim Return %"].notna()]
    if not lb_with_data.empty:
        avg_ret = lb_with_data["LB Sim Return %"].mean()
        winners = (lb_with_data["LB Sim Return %"] > 0).sum()
        print(f"  LOOKBACK ({LOOKBACK_DAYS}d):  "
              f"{len(strong_entries)} had strong DM entry signal | "
              f"{len(never_crossed)} never crossed 65 | "
              f"avg simulated return {avg_ret:+.1f}% | "
              f"{winners}/{len(lb_with_data)} positive")
    else:
        print(f"  LOOKBACK: no history data available")
    print("=" * 75)
    print()


# ============================================================
# EXCEL OUTPUT
# ============================================================

SIGNAL_COLORS = {
    "BACKED":    "C6EFCE",   # Green
    "WATCH":     "FFEB9C",   # Yellow
    "HEAD FAKE": "FFC7CE",   # Red
    "NO DATA":   "D9D9D9",   # Grey
}


def autofit_worksheet(ws):
    for col_cells in ws.columns:
        max_len    = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            value   = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 38)


def style_workbook(path: Path, row_count: int):
    wb    = load_workbook(path)
    ws    = wb["Top Gainers"]
    notes = wb["Notes"]

    ws["A1"] = "Top US Stock Gainers -- DM Cross-Reference + Lookback"
    ws["A2"] = "Generated:"
    ws["B2"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["A3"] = "DM Thresholds:"
    ws["B3"] = (f"BACKED >={DM_STRONG} | WATCH {DM_NEUTRAL}-{DM_STRONG} | "
                f"HEAD FAKE <{DM_NEUTRAL} | Lookback: {LOOKBACK_DAYS} days")

    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"].font = Font(bold=True)
    ws["A3"].font = Font(bold=True)

    header_fill    = PatternFill("solid", fgColor="1F4E79")   # dark blue -- current signal cols
    lb_header_fill = PatternFill("solid", fgColor="2E4057")   # darker blue -- lookback cols
    header_font    = Font(color="FFFFFF", bold=True)
    thin_border    = Border(
        left=Side(style="thin",   color="D9D9D9"),
        right=Side(style="thin",  color="D9D9D9"),
        top=Side(style="thin",    color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    header_row = 5
    last_row   = header_row + row_count
    last_col   = ws.max_column

    col_map = {cell.value: cell.column for cell in ws[header_row]}

    for cell in ws[header_row]:
        is_lb      = str(cell.value or "").startswith("LB ")
        cell.fill  = lb_header_fill if is_lb else header_fill
        cell.font  = header_font
        cell.border = thin_border

    signal_col = col_map.get("Signal")
    lb_ret_col = col_map.get("LB Sim Return %")

    pct_cols   = ["Today Change %"]
    money_cols = ["Price", "Today Change $", "Day Open", "Day High", "Day Low",
                  "Prev Close", "Last Trade", "Minute VWAP", "LB Entry Price"]
    vol_cols   = ["Day Volume"]

    for row_num in range(header_row + 1, last_row + 1):
        for col_num in range(1, last_col + 1):
            ws.cell(row_num, col_num).border = thin_border

        # Row color by current signal
        if signal_col:
            sv = ws.cell(row_num, signal_col).value or ""
            fc = SIGNAL_COLORS.get(sv)
            if fc:
                row_fill = PatternFill("solid", fgColor=fc)
                for col_num in range(1, last_col + 1):
                    ws.cell(row_num, col_num).fill = row_fill

        # Lookback return cell -- green/red by positive/negative
        if lb_ret_col:
            v = ws.cell(row_num, lb_ret_col).value
            if v is not None:
                try:
                    vf = float(v)
                    ws.cell(row_num, lb_ret_col).fill = PatternFill(
                        "solid", fgColor="C6EFCE" if vf >= 0 else "FFC7CE")
                    ws.cell(row_num, lb_ret_col).number_format = "+0.0%;-0.0%"
                except (ValueError, TypeError):
                    pass

        for name in pct_cols:
            if name in col_map:
                ws.cell(row_num, col_map[name]).number_format = "0.00%"
        for name in money_cols:
            if name in col_map:
                ws.cell(row_num, col_map[name]).number_format = "$0.00"
        for name in vol_cols:
            if name in col_map:
                ws.cell(row_num, col_map[name]).number_format = "#,##0"

    ws.freeze_panes = "A6"
    table_ref = f"A5:{get_column_letter(last_col)}{last_row}"
    table     = Table(displayName="TopGainersDMTable", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False, showLastColumn=False,
        showRowStripes=False, showColumnStripes=False,
    )
    ws.add_table(table)
    autofit_worksheet(ws)
    autofit_worksheet(notes)
    wb.save(path)


def write_dated_excel(df: pd.DataFrame) -> Path:
    OUTPUT_DIR.mkdir(exist_ok=True)
    stamp       = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    output_path = OUTPUT_DIR / f"top_gainers_dm_{stamp}.xlsx"

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Top Gainers", startrow=4, index=False)
        notes_data = pd.DataFrame({"Notes": [
            "Polygon market movers cross-referenced against DM_Latest.",
            f"BACKED = DM >= {DM_STRONG}: institutional flow confirmed, real move.",
            f"WATCH  = DM {DM_NEUTRAL}-{DM_STRONG}: mixed signal, monitor only.",
            f"HEAD FAKE = DM < {DM_NEUTRAL}: price up but institutions leaving.",
            "NO DATA = ticker not in DM universe.",
            "DM Source: Supabase dm_latest + dm_history tables",
            "",
            "LOOKBACK COLUMNS (LB prefix):",
            f"  Lookback source: Supabase dm_history",
            f"  Lookback window: {LOOKBACK_DAYS} trading days",
            "  LB Signal:",
            "    STRONG_ENTRY    = DM crossed 65+ at some point in the lookback window.",
            "    NEVER_CROSSED   = DM never reached 65 in the lookback window.",
            "    NOT_IN_HISTORY  = ticker not found in history file.",
            "    NO_HISTORY      = history file not available.",
            "  LB Entry Date:    first date in window where DM >= 65.",
            "  LB Entry Price:   closing price on that date.",
            "  LB Sim Return %:  (today price - entry price) / entry price.",
            "                    Positive = system would have profited.",
            "                    Negative = system would have lost.",
            "  LB Days Held:     calendar days from entry date to today.",
            "  LB Peak DM:       highest DM recorded in the lookback window.",
            "  LB Days Above 65: number of sessions DM >= 65 in window.",
            "",
            "LOOKBACK: Powered by Supabase dm_history table.",
        ]})
        notes_data.to_excel(writer, sheet_name="Notes", index=False)

    style_workbook(output_path, len(df))
    return output_path


def append_run_log(df: pd.DataFrame):
    OUTPUT_DIR.mkdir(exist_ok=True)
    log_path = OUTPUT_DIR / "top_gainers_history.csv"
    mode     = "a" if log_path.exists() else "w"
    header   = not log_path.exists()
    df.to_csv(log_path, mode=mode, header=header, index=False)


# ============================================================
# MAIN
# ============================================================

def fetch_dm_universe_movers(limit: int = TOP_N) -> pd.DataFrame:
    """
    Fetch snapshots for all DM universe tickers and find today's top movers
    with DM cross-reference + lookback analysis.
    """
    if not API_KEY:
        print("[WARN] No POLYGON_API_KEY -- cannot fetch universe movers.")
        return pd.DataFrame()

    dm_data = load_dm_data()
    if not dm_data:
        return pd.DataFrame()

    history_df = load_dm_history()
    dm_tickers = list(dm_data.keys())
    print(f"\n[DM UNIVERSE MOVERS] Checking {len(dm_tickers)} DM tickers...")

    all_snapshots = []
    batch_size = 50
    for i in range(0, len(dm_tickers), batch_size):
        batch = dm_tickers[i:i+batch_size]
        tickers_str = ",".join(batch)
        try:
            r = requests.get(
                "https://api.polygon.io/v2/snapshot/locale/us/markets/stocks/tickers",
                params={"apiKey": API_KEY, "tickers": tickers_str},
                timeout=30
            )
            if r.status_code == 200:
                all_snapshots.extend(r.json().get("tickers", []))
        except Exception:
            pass

    if not all_snapshots:
        print("  No snapshot data returned.")
        return pd.DataFrame()

    run_ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    rows = []
    for item in all_snapshots:
        day      = item.get("day", {}) or {}
        prev_day = item.get("prevDay", {}) or {}
        symbol   = item.get("ticker", "")
        chg_pct  = item.get("todaysChangePerc")

        if chg_pct is None or day.get("c") is None or day.get("c") == 0:
            continue

        current_price = day.get("c")
        dm_row   = dm_data.get(symbol.upper(), {})
        dm_score = dm_row.get("dm")
        signal   = classify_dm(dm_score)

        # Lookback analysis
        entry = find_entry_signal(history_df, symbol.upper(),
                                  current_price, LOOKBACK_DAYS) if history_df is not None else {
            "entry_date": None, "entry_dm": None, "entry_price": None,
            "days_held": None, "simulated_return": None,
            "signal_type": "NO_HISTORY", "peak_dm": None, "days_above_gate": None
        }

        rows.append({
            "Rank":              0,
            "Symbol":            symbol,
            "Signal":            signal,
            "DM Score":          dm_score,
            "DM Change":         dm_row.get("dm_chg"),
            "DM Phase":          dm_row.get("phase", ""),
            "Price":             current_price,
            "Today Change %":    float(chg_pct) / 100.0,
            "Today Change $":    item.get("todaysChange"),
            "Day Open":          day.get("o"),
            "Day High":          day.get("h"),
            "Day Low":           day.get("l"),
            "Day Volume":        day.get("v"),
            "Prev Close":        prev_day.get("c"),
            "Last Trade":        None,
            "Minute VWAP":       None,
            "LB Signal":         entry["signal_type"],
            "LB Entry Date":     entry["entry_date"],
            "LB Entry DM":       entry["entry_dm"],
            "LB Entry Price":    entry["entry_price"],
            "LB Days Held":      entry["days_held"],
            "LB Sim Return %":   entry["simulated_return"],
            "LB Peak DM":        entry["peak_dm"],
            "LB Days Above 65":  entry["days_above_gate"],
            "Retrieved At":      run_ts,
        })

    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows)
    df = df.sort_values("Today Change %", ascending=False).head(limit).reset_index(drop=True)
    df["Rank"] = range(1, len(df) + 1)
    print(f"  Found {len(df)} DM universe movers")
    return df


def main():
    # Part 1: Polygon top gainers (all stocks)
    print("=" * 75)
    print("  PART 1: POLYGON TOP GAINERS (all stocks)")
    print("=" * 75)
    df = fetch_top_gainers(limit=TOP_N)
    print_summary(df)
    xlsx_path = write_dated_excel(df)
    append_run_log(df)
    print(f"Excel: {xlsx_path.resolve()}")
    print(f"Log:   {(OUTPUT_DIR / 'top_gainers_history.csv').resolve()}")

    # Part 2: DM Universe top movers with lookback
    print()
    print("=" * 75)
    print("  PART 2: DM UNIVERSE TOP MOVERS + LOOKBACK")
    print("=" * 75)
    df_dm = fetch_dm_universe_movers(limit=TOP_N)
    if not df_dm.empty:
        print_summary(df_dm)
        OUTPUT_DIR.mkdir(exist_ok=True)
        stamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        dm_path = OUTPUT_DIR / f"dm_universe_movers_lookback_{stamp}.csv"
        df_dm.to_csv(dm_path, index=False)
        print(f"DM Movers CSV: {dm_path.resolve()}")
    else:
        print("No DM universe mover data available.")


if __name__ == "__main__":
    main()
