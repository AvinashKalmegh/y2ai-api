"""
Nitin_TickerETF_Step1and2.py
============================

Pipeline Step 1 + Step 2 of the Capital Flow methodology — Nitin engagement.

Step 1 — Polygon price/volume backfill for the 740 IUSV tickers + 17 anchor ETFs.
Step 2 — ETF correlation gate validation. Validates default ETF assignments or
          surfaces remap candidates for tickers that fail the gate.

Usage:
    python Nitin_TickerETF_Step1and2.py \
        --input  Nitin_TickerETF_Step1_Intake_2026-05-21.xlsx \
        --output Nitin_TickerETF_Step2_<date>.xlsx \
        --report Nitin_Step2_Report_<date>.md

Outputs:
    1. Updated Excel file (Tickers tab with Correlation, Status, Final_ETF, Notes filled).
    2. Markdown summary report (validated count, remap list, insufficient-data flags).
    3. Supabase tables populated with price/volume history (Avinash to adjust field names).

Prepared by: Vikram (specification) + Claude (starter code)
For: Avinash to adjust to his environment and execute
Date: May 21, 2026
"""

import argparse
import sys
import time
from datetime import datetime, timedelta
from pathlib import Path

import numpy as np
import pandas as pd
import requests
from openpyxl import load_workbook


# =============================================================================
# CONFIGURATION — ADJUST THESE FOR YOUR ENVIRONMENT
# =============================================================================

# Polygon API
POLYGON_API_KEY = "YOUR_POLYGON_KEY_HERE"  # TODO: pull from env var or config
POLYGON_BASE_URL = "https://api.polygon.io"

# Date range for backfill
BACKFILL_START_DATE = "2016-01-01"
BACKFILL_END_DATE = datetime.now().strftime("%Y-%m-%d")

# Supabase configuration — TODO: adjust to your existing pipeline
# Avinash: these are placeholders. Use your existing dm_history-style schema
# and adjust the table/column names. Schema suggestion below.
SUPABASE_URL = "YOUR_SUPABASE_URL"
SUPABASE_KEY = "YOUR_SUPABASE_KEY"

# Shared price/volume tables — all clients read from the same raw data layer.
# If a ticker's history is already loaded for any prior engagement, we reuse it
# rather than re-pulling from Polygon. Architecture: shared raw data, per-client
# analytical layer.
SUPABASE_PRICE_TABLE = "price_history"      # the existing 597-ticker price table
SUPABASE_ETF_TABLE = "etf_history"          # the existing ETF price table
SUPABASE_DM_TABLE = "dm_history"            # the existing CF/DM history table

# =============================================================================
# METHODOLOGY CONSTANTS — DO NOT CHANGE
# These match the operational methodology. Changes here break consistency
# with our own universe and require explicit decision + documentation update.
# =============================================================================

CORRELATION_WINDOW = 252       # trading days — aligns with CF formula window
GATE_THRESHOLD = 0.35          # methodology Tier B correlation gate
MIN_DAYS_FOR_VALID_GATE = 252  # below this, ticker is flagged as insufficient
REMAP_CANDIDATES = ["IGV", "SMH", "ITA", "XBI", "IUSV", "IWD"]


# =============================================================================
# POLYGON DATA INGESTION
# =============================================================================

def fetch_polygon_daily(ticker: str, start: str, end: str) -> pd.DataFrame:
    """
    Pull daily OHLCV from Polygon aggregates endpoint.
    Returns a DataFrame with columns: date, open, high, low, close, volume.
    """
    url = f"{POLYGON_BASE_URL}/v2/aggs/ticker/{ticker}/range/1/day/{start}/{end}"
    params = {"adjusted": "true", "sort": "asc", "limit": 50000, "apiKey": POLYGON_API_KEY}

    try:
        r = requests.get(url, params=params, timeout=30)
        r.raise_for_status()
        data = r.json()
    except requests.RequestException as e:
        print(f"  [WARN] {ticker}: Polygon request failed — {e}")
        return pd.DataFrame()

    if data.get("status") != "OK" or not data.get("results"):
        print(f"  [WARN] {ticker}: Polygon returned no results (status={data.get('status')})")
        return pd.DataFrame()

    df = pd.DataFrame(data["results"])
    df["date"] = pd.to_datetime(df["t"], unit="ms").dt.date
    df = df.rename(columns={"o":"open","h":"high","l":"low","c":"close","v":"volume"})
    return df[["date","open","high","low","close","volume"]].sort_values("date").reset_index(drop=True)


def get_existing_symbols_from_supabase() -> set:
    """
    Query Supabase for the set of symbols (tickers + ETFs) that already have
    price history loaded from prior engagements or our own universe.

    AVINASH: replace with your supabase-py select pattern.
    Suggested query (across both price_history and etf_history tables):
        SELECT DISTINCT ticker FROM price_history
        UNION
        SELECT DISTINCT etf AS ticker FROM etf_history

    Returns a set of symbol strings.
    """
    print("\n[OVERLAP] Querying Supabase for existing symbols...")
    print("[OVERLAP] STUB — Avinash: replace with your supabase-py select")
    # Example shape:
    # price_rows = supabase.table(SUPABASE_PRICE_TABLE).select("ticker").execute()
    # etf_rows   = supabase.table(SUPABASE_ETF_TABLE).select("etf").execute()
    # existing = {r["ticker"] for r in price_rows.data} | {r["etf"] for r in etf_rows.data}
    # return existing

    # Until the stub is wired up, return empty set — script will pull everything
    # from Polygon. After Avinash wires this in, the overlap is detected and
    # we skip duplicates.
    return set()


def load_from_supabase(symbol: str, is_etf: bool = False) -> pd.DataFrame:
    """
    Load price/volume history for a single symbol from Supabase.

    AVINASH: replace with your supabase-py select pattern.
    Expected output: DataFrame with columns [date, open, high, low, close, volume]
    matching the shape returned by fetch_polygon_daily().
    """
    # Example shape:
    # table = SUPABASE_ETF_TABLE if is_etf else SUPABASE_PRICE_TABLE
    # col = "etf" if is_etf else "ticker"
    # rows = supabase.table(table).select("*").eq(col, symbol).order("date").execute()
    # df = pd.DataFrame(rows.data)
    # df["date"] = pd.to_datetime(df["date"]).dt.date
    # return df[["date","open","high","low","close","volume"]]
    raise NotImplementedError("Wire up to supabase-py select")


def backfill_all(tickers: list, etfs: list, throttle_seconds: float = 0.15) -> tuple:
    """
    Pull price history for every ticker and ETF, with overlap optimization.

    Step 1: query Supabase for symbols we already have. Reuse those.
    Step 2: for symbols not yet loaded, pull from Polygon and write to Supabase.
    Step 3: return (data dict, newly_pulled set) for downstream use.
    """
    all_symbols = list(dict.fromkeys(tickers + etfs))   # dedupe, preserve order
    print(f"\n[BACKFILL] {len(all_symbols)} symbols requested ({len(tickers)} tickers + {len(etfs)} ETFs)")

    # Overlap check
    existing = get_existing_symbols_from_supabase()
    reuse = [s for s in all_symbols if s in existing]
    to_pull = [s for s in all_symbols if s not in existing]
    print(f"[BACKFILL] Reusing from Supabase: {len(reuse)} symbols")
    print(f"[BACKFILL] Pulling from Polygon:  {len(to_pull)} symbols")
    print(f"[BACKFILL] Date range: {BACKFILL_START_DATE} to {BACKFILL_END_DATE}")

    data = {}
    newly_pulled = set()

    # Load existing symbols from Supabase
    for sym in reuse:
        try:
            data[sym] = load_from_supabase(sym, is_etf=(sym in etfs))
        except NotImplementedError:
            # Stub not yet wired — fall through to Polygon
            to_pull.append(sym)

    # Pull new symbols from Polygon
    failures = []
    for i, sym in enumerate(to_pull, 1):
        if i % 50 == 0:
            print(f"  [{i}/{len(to_pull)}] {sym}")
        df = fetch_polygon_daily(sym, BACKFILL_START_DATE, BACKFILL_END_DATE)
        if not df.empty:
            data[sym] = df
            newly_pulled.add(sym)
        else:
            failures.append(sym)
        time.sleep(throttle_seconds)

    print(f"[BACKFILL] Complete. {len(data)} symbols loaded ({len(newly_pulled)} new from Polygon, {len(data) - len(newly_pulled)} reused), {len(failures)} failed.")
    if failures:
        print(f"[BACKFILL] Failures: {failures[:20]}{' ...' if len(failures) > 20 else ''}")
    return data, newly_pulled


def write_to_supabase(data: dict, tickers: list, etfs: list, newly_pulled: set):
    """
    Push newly-pulled price/volume history into Supabase.

    Only writes symbols that were just pulled from Polygon (newly_pulled).
    Symbols loaded from Supabase via load_from_supabase() are not re-written.

    AVINASH: adjust this to your existing pipeline conventions.
    The shape we need on the other side is:
      - one table for ticker prices (long format: ticker, date, ohlcv)
      - one table for ETF prices  (same shape, separate table for clarity)

    Schema (matches your existing dm_history-style pattern):
      CREATE TABLE price_history (
          ticker  TEXT NOT NULL,
          date    DATE NOT NULL,
          open    NUMERIC, high NUMERIC, low NUMERIC, close NUMERIC,
          volume  BIGINT,
          PRIMARY KEY (ticker, date)
      );
      CREATE INDEX ON price_history (date);
      CREATE INDEX ON price_history (ticker);

    Same for etf_history with etf in place of ticker.

    These tables are shared across all client engagements. Each engagement
    only adds tickers that aren't already there. The methodology layer
    (Tier A, Tier B, walk-forward) lives in per-client tables but the
    raw price/volume data is global.

    Stub implementation below — replace with your supabase-py upsert pattern.
    """
    to_write = {sym: df for sym, df in data.items() if sym in newly_pulled}
    print(f"\n[SUPABASE] Writing {len(to_write)} newly-pulled symbols to shared tables")
    print(f"[SUPABASE] ({len(data) - len(to_write)} symbols already in Supabase, no write needed)")
    print("[SUPABASE] STUB — Avinash: replace with your supabase-py upsert pattern")
    # Example shape:
    # for symbol, df in to_write.items():
    #     table = SUPABASE_ETF_TABLE if symbol in etfs else SUPABASE_PRICE_TABLE
    #     col = "etf" if symbol in etfs else "ticker"
    #     records = df.assign(**{col: symbol}).to_dict(orient="records")
    #     supabase.table(table).upsert(records, on_conflict=f"{col},date").execute()


# =============================================================================
# CORRELATION GATE — STEP 2
# =============================================================================

def compute_correlation(ticker_df: pd.DataFrame, etf_df: pd.DataFrame,
                        window: int = CORRELATION_WINDOW) -> tuple:
    """
    Compute Pearson correlation between ticker and ETF on daily log returns
    over the trailing `window` trading days.

    Returns (correlation, days_used). If insufficient overlapping data,
    returns (None, days_available).
    """
    # Merge on date to get aligned series
    merged = pd.merge(
        ticker_df[["date", "close"]].rename(columns={"close": "ticker_close"}),
        etf_df[["date", "close"]].rename(columns={"close": "etf_close"}),
        on="date", how="inner"
    ).sort_values("date").reset_index(drop=True)

    if len(merged) < 2:
        return (None, 0)

    # Log returns
    merged["ticker_ret"] = np.log(merged["ticker_close"] / merged["ticker_close"].shift(1))
    merged["etf_ret"] = np.log(merged["etf_close"] / merged["etf_close"].shift(1))
    merged = merged.dropna()

    if len(merged) == 0:
        return (None, 0)

    # Take the trailing window
    window_data = merged.tail(window)
    days_used = len(window_data)

    if days_used < 2:
        return (None, days_used)

    corr = window_data["ticker_ret"].corr(window_data["etf_ret"])
    return (float(corr), days_used)


def validate_assignments(tickers_df: pd.DataFrame, data: dict) -> pd.DataFrame:
    """
    For each ticker, compute correlation against its Default_ETF.
    Pass = correlation >= GATE_THRESHOLD with sufficient data window.
    Fail = compute correlation against all REMAP_CANDIDATES, pick best alternative.

    Writes Correlation, Status, Final_ETF, Notes columns into the DataFrame.
    """
    print(f"\n[GATE] Running correlation validation over trailing {CORRELATION_WINDOW} days")
    print(f"[GATE] Threshold: {GATE_THRESHOLD}")

    out = tickers_df.copy()
    out["Correlation"] = np.nan
    out["Status"] = "PENDING"
    out["Final_ETF"] = ""
    out["Notes"] = ""

    for idx, row in out.iterrows():
        ticker = row["Ticker"]
        default_etf = row["Default_ETF"]

        if ticker not in data or default_etf not in data:
            out.at[idx, "Status"] = "MISSING_DATA"
            out.at[idx, "Notes"] = f"Polygon data unavailable for ticker or default ETF ({default_etf})"
            continue

        corr, days = compute_correlation(data[ticker], data[default_etf])

        if corr is None:
            out.at[idx, "Status"] = "MISSING_DATA"
            out.at[idx, "Notes"] = f"Insufficient overlapping data (days_used={days})"
            continue

        out.at[idx, "Correlation"] = round(corr, 4)

        if days < MIN_DAYS_FOR_VALID_GATE:
            out.at[idx, "Status"] = "INSUFFICIENT_HISTORY"
            out.at[idx, "Notes"] = f"Only {days} days of history available; correlation reported but gate not fully validated"
            out.at[idx, "Final_ETF"] = default_etf
            continue

        if corr >= GATE_THRESHOLD:
            out.at[idx, "Status"] = "VALIDATED"
            out.at[idx, "Final_ETF"] = default_etf
        else:
            # Below gate — test remap candidates
            best_alt = None
            best_corr = -1.0
            alt_results = []
            for alt_etf in REMAP_CANDIDATES:
                if alt_etf == default_etf or alt_etf not in data:
                    continue
                alt_corr, alt_days = compute_correlation(data[ticker], data[alt_etf])
                if alt_corr is not None and alt_days >= MIN_DAYS_FOR_VALID_GATE:
                    alt_results.append(f"{alt_etf}:{alt_corr:.3f}")
                    if alt_corr > best_corr:
                        best_corr = alt_corr
                        best_alt = alt_etf

            out.at[idx, "Status"] = "REMAPPED_NEEDED"
            if best_alt is not None and best_corr >= GATE_THRESHOLD:
                out.at[idx, "Notes"] = (
                    f"Default {default_etf} corr={corr:.3f} below gate. "
                    f"Best alternative: {best_alt} (corr={best_corr:.3f}). "
                    f"All alternatives tested: {', '.join(alt_results)}"
                )
            else:
                out.at[idx, "Notes"] = (
                    f"Default {default_etf} corr={corr:.3f} below gate. "
                    f"No alternative cleared {GATE_THRESHOLD}. "
                    f"Best alternative tested: {best_alt} (corr={best_corr:.3f}). "
                    f"Manual review required."
                )

    # Summary counts
    counts = out["Status"].value_counts()
    print(f"\n[GATE] Summary:")
    for status, count in counts.items():
        print(f"  {status}: {count}")

    return out


# =============================================================================
# OUTPUT WRITERS
# =============================================================================

def write_excel_output(input_path: str, output_path: str, results_df: pd.DataFrame):
    """
    Copy the input workbook to output_path and update the Tickers tab in place
    with the correlation results. Preserves all formatting and other tabs.
    """
    print(f"\n[OUTPUT] Writing updated Excel to: {output_path}")
    wb = load_workbook(input_path)
    ws = wb["Tickers"]

    # Find column indices by header
    headers = {ws.cell(row=4, column=c).value: c for c in range(1, ws.max_column + 1)}
    required = ["Ticker", "Status", "Final_ETF", "Correlation", "Notes"]
    for col in required:
        if col not in headers:
            raise ValueError(f"Expected column '{col}' not found in Tickers tab")

    # Build a ticker → result lookup
    results_lookup = results_df.set_index("Ticker").to_dict("index")

    # Write back from row 5 onward
    for row_idx in range(5, ws.max_row + 1):
        ticker = ws.cell(row=row_idx, column=headers["Ticker"]).value
        if ticker not in results_lookup:
            continue
        r = results_lookup[ticker]
        ws.cell(row=row_idx, column=headers["Status"], value=r["Status"])
        ws.cell(row=row_idx, column=headers["Final_ETF"], value=r["Final_ETF"])
        corr_val = r["Correlation"]
        if pd.notna(corr_val):
            ws.cell(row=row_idx, column=headers["Correlation"], value=float(corr_val))
        ws.cell(row=row_idx, column=headers["Notes"], value=r["Notes"])

    wb.save(output_path)
    print(f"[OUTPUT] Saved.")


def write_summary_report(report_path: str, results_df: pd.DataFrame):
    """
    Produce a markdown summary report for Vikram review.
    """
    print(f"\n[REPORT] Writing summary report to: {report_path}")

    validated = results_df[results_df["Status"] == "VALIDATED"]
    remap = results_df[results_df["Status"] == "REMAPPED_NEEDED"]
    insufficient = results_df[results_df["Status"] == "INSUFFICIENT_HISTORY"]
    missing = results_df[results_df["Status"] == "MISSING_DATA"]

    lines = []
    lines.append("# Nitin Engagement — Step 2 Correlation Gate Report")
    lines.append("")
    lines.append(f"**Run date:** {datetime.now().strftime('%Y-%m-%d %H:%M')}  ")
    lines.append(f"**Universe:** {len(results_df)} IUSV holdings  ")
    lines.append(f"**Correlation window:** {CORRELATION_WINDOW} trading days  ")
    lines.append(f"**Gate threshold:** {GATE_THRESHOLD}  ")
    lines.append("")
    lines.append("## Summary")
    lines.append("")
    lines.append(f"- VALIDATED: {len(validated)}")
    lines.append(f"- REMAPPED_NEEDED: {len(remap)}")
    lines.append(f"- INSUFFICIENT_HISTORY: {len(insufficient)}")
    lines.append(f"- MISSING_DATA: {len(missing)}")
    lines.append("")

    if len(remap) > 0:
        lines.append("## Tickers needing remap")
        lines.append("")
        lines.append("These tickers failed the gate against their default sector ETF. Best alternative anchor identified for each. Manual review required before promotion to Step 3.")
        lines.append("")
        lines.append("| Ticker | Name | Default ETF | Correlation | Notes |")
        lines.append("|---|---|---|---|---|")
        for _, r in remap.iterrows():
            name = (r["Name"] or "")[:40]
            corr = r["Correlation"] if pd.notna(r["Correlation"]) else "—"
            corr_str = f"{corr:.3f}" if isinstance(corr, float) else corr
            lines.append(f"| {r['Ticker']} | {name} | {r['Default_ETF']} | {corr_str} | {r['Notes']} |")
        lines.append("")

    if len(insufficient) > 0:
        lines.append("## Tickers with insufficient history")
        lines.append("")
        lines.append(f"Less than {CORRELATION_WINDOW} trading days of overlapping data with default ETF. Correlation reported on available data but gate not fully validated.")
        lines.append("")
        lines.append("| Ticker | Name | Default ETF | Correlation | Notes |")
        lines.append("|---|---|---|---|---|")
        for _, r in insufficient.iterrows():
            name = (r["Name"] or "")[:40]
            corr = r["Correlation"] if pd.notna(r["Correlation"]) else "—"
            corr_str = f"{corr:.3f}" if isinstance(corr, float) else corr
            lines.append(f"| {r['Ticker']} | {name} | {r['Default_ETF']} | {corr_str} | {r['Notes']} |")
        lines.append("")

    if len(missing) > 0:
        lines.append("## Tickers with missing data")
        lines.append("")
        lines.append("Polygon returned no data for these tickers. Possible delisting, ticker change, or API issue. Manual review.")
        lines.append("")
        for _, r in missing.iterrows():
            lines.append(f"- **{r['Ticker']}** ({r['Name']}): {r['Notes']}")
        lines.append("")

    lines.append("## Sample of validated tickers")
    lines.append("")
    lines.append("Top 20 validated by IUSV weight, for spot-check:")
    lines.append("")
    lines.append("| Ticker | Default ETF | Correlation |")
    lines.append("|---|---|---|")
    sample = validated.sort_values("IUSV_Weight_Pct", ascending=False).head(20)
    for _, r in sample.iterrows():
        lines.append(f"| {r['Ticker']} | {r['Default_ETF']} | {r['Correlation']:.3f} |")
    lines.append("")

    lines.append("## Next step")
    lines.append("")
    lines.append("Vikram reviews the REMAPPED_NEEDED list and confirms the alternative ETF assignment for each. Final_ETF column in the updated Excel gets set to the confirmed ETF (default or remapped). Once approved, the file becomes the input to Step 3 (CF computation).")
    lines.append("")

    Path(report_path).write_text("\n".join(lines), encoding="utf-8")
    print(f"[REPORT] Saved.")


# =============================================================================
# MAIN
# =============================================================================

def main():
    ap = argparse.ArgumentParser(description="Nitin engagement Step 1 + Step 2 pipeline")
    ap.add_argument("--input",  required=True, help="Path to intake Excel file")
    ap.add_argument("--output", required=True, help="Path to write updated Excel file")
    ap.add_argument("--report", required=True, help="Path to write markdown summary report")
    ap.add_argument("--skip-backfill", action="store_true",
                    help="Skip Polygon backfill and Supabase write — useful when re-running validation on existing data")
    args = ap.parse_args()

    # Load intake file
    print(f"[INTAKE] Reading: {args.input}")
    tickers_df = pd.read_excel(args.input, sheet_name="Tickers", skiprows=3)
    etfs_df = pd.read_excel(args.input, sheet_name="ETF_List", skiprows=3)

    tickers = tickers_df["Ticker"].dropna().tolist()
    etfs = etfs_df["ETF"].dropna().tolist()
    print(f"[INTAKE] {len(tickers)} tickers, {len(etfs)} ETFs")

    # Step 1: Polygon backfill (overlap-aware)
    if not args.skip_backfill:
        data, newly_pulled = backfill_all(tickers, etfs)
        write_to_supabase(data, tickers, etfs, newly_pulled)
    else:
        # Avinash: if re-running, load from Supabase instead of Polygon.
        # Replace with your supabase-py select pattern that returns
        # {symbol: DataFrame(date, open, high, low, close, volume)}
        print("[INTAKE] --skip-backfill set; loading from Supabase instead")
        raise NotImplementedError("Adjust this branch to load from Supabase")

    # Step 2: correlation gate validation
    results = validate_assignments(tickers_df, data)

    # Write outputs
    write_excel_output(args.input, args.output, results)
    write_summary_report(args.report, results)

    print("\n[DONE] Step 1 + Step 2 complete. Hand off to Vikram for review.")


if __name__ == "__main__":
    main()
