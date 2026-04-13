import os
from datetime import datetime
from pathlib import Path
from dotenv import load_dotenv

load_dotenv()

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ============================================================
# CONFIGURATION
# ============================================================

API_KEY = os.getenv("POLYGON_API_KEY", "")
API_URL = "https://api.polygon.io/v2/snapshot/locale/us/markets/stocks/gainers"
OUTPUT_DIR = Path("market_snapshots_polygon")
TOP_N = 20

# DM classification thresholds
DM_STRONG  = 65   # Institutionally backed -- enter candidate
DM_NEUTRAL = 40   # Mixed -- watch only
# Below DM_NEUTRAL = head fake territory

# Minimum price and volume to filter out OTC/micro-caps
MIN_PRICE  = 5.0
MIN_VOLUME = 100000


# ============================================================
# DM LOADER (from Supabase dm_latest table)
# ============================================================

def load_dm_data() -> dict:
    """
    Load DM data from Supabase dm_latest table.
    Returns dict: ticker -> {dm, dm_chg, phase, dm_close}
    """
    url = os.getenv("SUPABASE_URL")
    key = os.getenv("SUPABASE_KEY")
    if not url or not key:
        print("[WARN] SUPABASE_URL/KEY not set -- running without DM data.")
        return {}

    try:
        from supabase import create_client
        client = create_client(url, key)

        rows = []
        offset, page = 0, 5000
        while True:
            r = (client.table('dm_latest')
                 .select('ticker,dm_smoothed,dm_change,phase,close')
                 .range(offset, offset + page - 1)
                 .execute())
            rows.extend(r.data)
            if len(r.data) < page:
                break
            offset += page

        result = {}
        for row in rows:
            ticker = str(row['ticker']).strip().upper()
            result[ticker] = {
                "dm":       float(row['dm_smoothed']) if row.get('dm_smoothed') is not None else None,
                "dm_chg":   float(row['dm_change'])   if row.get('dm_change') is not None else None,
                "phase":    str(row.get('phase', '')),
                "dm_close": float(row['close'])       if row.get('close') is not None else None,
            }
        print(f"[DM] Loaded {len(result)} tickers from Supabase dm_latest")
        return result

    except Exception as e:
        print(f"[WARN] Failed to load DM data from Supabase: {e}")
        return {}


def classify_dm(dm_score) -> str:
    if dm_score is None:
        return "NO DATA"
    if dm_score >= DM_STRONG:
        return "BACKED"       # Institutional flow confirmed -- real move
    if dm_score >= DM_NEUTRAL:
        return "WATCH"        # Mixed -- monitor, don't chase
    return "HEAD FAKE"        # Price up, institutions leaving -- avoid


# ============================================================
# POLYGON FETCH
# ============================================================

def _pick_rows(payload):
    if isinstance(payload, list):
        return payload
    if isinstance(payload, dict):
        for key in ("tickers", "results"):
            if isinstance(payload.get(key), list):
                return payload[key]
    raise RuntimeError(f"Unexpected Polygon response shape: {payload}")


def fetch_top_gainers(limit: int = TOP_N) -> pd.DataFrame:
    if not API_KEY:
        raise RuntimeError(
            "Missing POLYGON_API_KEY environment variable. "
            "Set it first, then rerun the script."
        )

    resp = requests.get(API_URL, params={"apiKey": API_KEY}, timeout=30)
    resp.raise_for_status()
    payload = resp.json()
    rows = _pick_rows(payload)

    if not rows:
        raise RuntimeError("Polygon returned no gainers rows.")

    # Load DM data from Supabase
    dm_data = load_dm_data()

    # Filter out OTC/micro-caps and build results
    normalized = []
    run_ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    rank = 0

    for item in rows:
        day        = item.get("day", {}) or {}
        prev_day   = item.get("prevDay", {}) or {}
        last_trade = item.get("lastTrade", {}) or {}
        min_bar    = item.get("min", {}) or {}

        symbol     = item.get("ticker") or item.get("symbol")
        price      = day.get("c") or 0
        volume     = day.get("v") or 0
        change_pct = item.get("todaysChangePerc")

        # Filter: skip penny stocks and low-volume tickers
        if price < MIN_PRICE or volume < MIN_VOLUME:
            continue

        if change_pct is not None:
            change_pct = float(change_pct) / 100.0

        # DM enrichment
        dm_row    = dm_data.get(symbol.upper(), {}) if symbol else {}
        dm_score  = dm_row.get("dm")
        dm_change = dm_row.get("dm_chg")
        dm_phase  = dm_row.get("phase", "")
        signal    = classify_dm(dm_score)

        rank += 1
        normalized.append({
            "Rank":             rank,
            "Symbol":           symbol,
            "Signal":           signal,
            "DM Score":         dm_score,
            "DM Change":        dm_change,
            "DM Phase":         dm_phase,
            "Price":            price,
            "Today Change %":   change_pct,
            "Today Change $":   item.get("todaysChange"),
            "Day Open":         day.get("o"),
            "Day High":         day.get("h"),
            "Day Low":          day.get("l"),
            "Day Volume":       volume,
            "Prev Close":       prev_day.get("c"),
            "Last Trade":       last_trade.get("p"),
            "Minute VWAP":      min_bar.get("vw"),
            "Retrieved At":     run_ts,
        })

        if rank >= limit:
            break

    if not normalized:
        print("[WARN] No gainers passed the filter (min price $%.0f, min volume %d)" % (MIN_PRICE, MIN_VOLUME))
        print("       This may happen outside market hours.")

    return pd.DataFrame(normalized)


# ============================================================
# CONSOLE SUMMARY
# ============================================================

def print_summary(df: pd.DataFrame):
    print()
    print("=" * 65)
    print(f"  TOP GAINERS -- DM CROSS-REFERENCE  {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print("=" * 65)

    backed     = df[df["Signal"] == "BACKED"]
    watch      = df[df["Signal"] == "WATCH"]
    head_fakes = df[df["Signal"] == "HEAD FAKE"]
    no_data    = df[df["Signal"] == "NO DATA"]

    if not backed.empty:
        print(f"\n  INSTITUTIONALLY BACKED (DM >= {DM_STRONG}) -- {len(backed)} names:")
        print(f"    {'Symbol':<8} {'Change%':>8}  {'DM':>6}  {'DM Chg':>8}  Phase")
        print(f"    {'-'*7:<8} {'-'*7:>8}  {'-'*6:>6}  {'-'*8:>8}  {'-'*10}")
        for _, r in backed.iterrows():
            chg = f"{r['Today Change %']:.1%}" if pd.notna(r['Today Change %']) else "N/A"
            dm  = f"{r['DM Score']:.1f}" if pd.notna(r['DM Score']) else "--"
            dmc = f"{r['DM Change']:+.1f}" if pd.notna(r['DM Change']) else "--"
            print(f"    {r['Symbol']:<8} {chg:>8}  {dm:>6}  {dmc:>8}  {r['DM Phase']}")

    if not watch.empty:
        print(f"\n  WATCH -- MIXED SIGNAL (DM {DM_NEUTRAL}-{DM_STRONG}) -- {len(watch)} names:")
        print(f"    {'Symbol':<8} {'Change%':>8}  {'DM':>6}  {'DM Chg':>8}  Phase")
        print(f"    {'-'*7:<8} {'-'*7:>8}  {'-'*6:>6}  {'-'*8:>8}  {'-'*10}")
        for _, r in watch.iterrows():
            chg = f"{r['Today Change %']:.1%}" if pd.notna(r['Today Change %']) else "N/A"
            dm  = f"{r['DM Score']:.1f}" if pd.notna(r['DM Score']) else "--"
            dmc = f"{r['DM Change']:+.1f}" if pd.notna(r['DM Change']) else "--"
            print(f"    {r['Symbol']:<8} {chg:>8}  {dm:>6}  {dmc:>8}  {r['DM Phase']}")

    if not head_fakes.empty:
        print(f"\n  HEAD FAKES (DM < {DM_NEUTRAL}) -- DO NOT CHASE -- {len(head_fakes)} names:")
        print(f"    {'Symbol':<8} {'Change%':>8}  {'DM':>6}  {'DM Chg':>8}  Phase")
        print(f"    {'-'*7:<8} {'-'*7:>8}  {'-'*6:>6}  {'-'*8:>8}  {'-'*10}")
        for _, r in head_fakes.iterrows():
            chg = f"{r['Today Change %']:.1%}" if pd.notna(r['Today Change %']) else "N/A"
            dm  = f"{r['DM Score']:.1f}" if pd.notna(r['DM Score']) else "--"
            dmc = f"{r['DM Change']:+.1f}" if pd.notna(r['DM Change']) else "--"
            print(f"    {r['Symbol']:<8} {chg:>8}  {dm:>6}  {dmc:>8}  {r['DM Phase']}")

    if not no_data.empty:
        tickers = ", ".join(no_data["Symbol"].tolist())
        print(f"\n  NOT IN DM UNIVERSE ({len(no_data)}): {tickers}")

    print()
    print(f"  SUMMARY: {len(backed)} backed | {len(watch)} watch | {len(head_fakes)} head fakes | {len(no_data)} no data")
    print("=" * 65)
    print()


# ============================================================
# EXCEL OUTPUT
# ============================================================

SIGNAL_COLORS = {
    "BACKED":    "C6EFCE",   # Green
    "WATCH":     "FFEB9C",   # Yellow
    "HEAD FAKE": "FFC7CE",   # Red
    "NO DATA":   "D9D9D9",   # Grey
}


def autofit_worksheet(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 38)


def style_workbook(path: Path, row_count: int):
    wb = load_workbook(path)
    ws = wb["Top Gainers"]
    notes = wb["Notes"]

    ws["A1"] = "Top US Stock Gainers -- DM Cross-Reference"
    ws["A2"] = "Generated:"
    ws["B2"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["A3"] = "DM Thresholds:"
    ws["B3"] = f"BACKED >={DM_STRONG} | WATCH {DM_NEUTRAL}-{DM_STRONG} | HEAD FAKE <{DM_NEUTRAL}"

    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"].font = Font(bold=True)
    ws["A3"].font = Font(bold=True)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    header_row = 5
    last_row   = header_row + row_count
    last_col   = ws.max_column

    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    col_map = {cell.value: cell.column for cell in ws[header_row]}
    signal_col = col_map.get("Signal")

    for row_num in range(header_row + 1, last_row + 1):
        for col_num in range(1, last_col + 1):
            ws.cell(row_num, col_num).border = thin_border

        # Colour entire row by signal
        if signal_col:
            signal_val = ws.cell(row_num, signal_col).value or ""
            fill_color = SIGNAL_COLORS.get(signal_val)
            if fill_color:
                row_fill = PatternFill("solid", fgColor=fill_color)
                for col_num in range(1, last_col + 1):
                    ws.cell(row_num, col_num).fill = row_fill

    # Number formats
    pct_cols   = ["Today Change %"]
    money_cols = ["Price", "Today Change $", "Day Open", "Day High", "Day Low",
                  "Prev Close", "Last Trade", "Minute VWAP"]
    vol_cols   = ["Day Volume"]

    for row_num in range(header_row + 1, last_row + 1):
        for name in pct_cols:
            if name in col_map:
                ws.cell(row_num, col_map[name]).number_format = "0.00%"
        for name in money_cols:
            if name in col_map:
                ws.cell(row_num, col_map[name]).number_format = "$0.00"
        for name in vol_cols:
            if name in col_map:
                ws.cell(row_num, col_map[name]).number_format = "#,##0"

    ws.freeze_panes = "A6"
    table_ref = f"A5:{get_column_letter(last_col)}{last_row}"
    table = Table(displayName="TopGainersDMTable", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,   # Off -- we use signal colours instead
        showColumnStripes=False,
    )
    ws.add_table(table)

    autofit_worksheet(ws)
    autofit_worksheet(notes)
    wb.save(path)


def write_dated_excel(df: pd.DataFrame) -> Path:
    OUTPUT_DIR.mkdir(exist_ok=True)
    stamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    output_path = OUTPUT_DIR / f"top_gainers_dm_{stamp}.xlsx"

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Top Gainers", startrow=4, index=False)
        notes_data = pd.DataFrame({
            "Notes": [
                "Polygon market movers cross-referenced against DM_Latest.",
                f"BACKED = DM >= {DM_STRONG}: institutional flow confirmed, real move.",
                f"WATCH  = DM {DM_NEUTRAL}-{DM_STRONG}: mixed signal, monitor only.",
                f"HEAD FAKE = DM < {DM_NEUTRAL}: price up but institutions leaving -- do not chase.",
                "NO DATA = ticker not in DM universe.",
                "DM Source: Supabase dm_latest table",
                f"Filters: price >= ${MIN_PRICE}, volume >= {MIN_VOLUME:,}",
            ]
        })
        notes_data.to_excel(writer, sheet_name="Notes", index=False)

    style_workbook(output_path, len(df))
    return output_path


def append_run_log(df: pd.DataFrame):
    OUTPUT_DIR.mkdir(exist_ok=True)
    log_path = OUTPUT_DIR / "top_gainers_history.csv"
    mode   = "a" if log_path.exists() else "w"
    header = not log_path.exists()
    df.to_csv(log_path, mode=mode, header=header, index=False)


# ============================================================
# MAIN
# ============================================================

def fetch_dm_universe_movers() -> pd.DataFrame:
    """
    Fetch snapshots for all DM universe tickers and find today's top movers.
    This catches large-cap gainers that Polygon's top-20 misses.
    """
    dm_data = load_dm_data()
    if not dm_data:
        return pd.DataFrame()

    dm_tickers = list(dm_data.keys())
    print(f"\n[DM UNIVERSE MOVERS] Checking {len(dm_tickers)} DM tickers...")

    # Fetch snapshots in batches (Polygon allows comma-separated tickers)
    all_snapshots = []
    batch_size = 50
    for i in range(0, len(dm_tickers), batch_size):
        batch = dm_tickers[i:i+batch_size]
        tickers_str = ",".join(batch)
        try:
            r = requests.get(
                "https://api.polygon.io/v2/snapshot/locale/us/markets/stocks/tickers",
                params={"apiKey": API_KEY, "tickers": tickers_str},
                timeout=30
            )
            if r.status_code == 200:
                data = r.json()
                all_snapshots.extend(data.get("tickers", []))
        except Exception:
            pass

    if not all_snapshots:
        print("  No snapshot data returned.")
        return pd.DataFrame()

    # Build results sorted by change %
    run_ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    rows = []
    for item in all_snapshots:
        day      = item.get("day", {}) or {}
        prev_day = item.get("prevDay", {}) or {}
        symbol   = item.get("ticker", "")
        chg_pct  = item.get("todaysChangePerc")

        if chg_pct is None or day.get("c") is None or day.get("c") == 0:
            continue

        dm_row   = dm_data.get(symbol.upper(), {})
        dm_score = dm_row.get("dm")
        signal   = classify_dm(dm_score)

        rows.append({
            "Rank":           0,
            "Symbol":         symbol,
            "Signal":         signal,
            "DM Score":       dm_score,
            "DM Change":      dm_row.get("dm_chg"),
            "DM Phase":       dm_row.get("phase", ""),
            "Price":          day.get("c"),
            "Today Change %": float(chg_pct) / 100.0,
            "Today Change $": item.get("todaysChange"),
            "Day Open":       day.get("o"),
            "Day High":       day.get("h"),
            "Day Low":        day.get("l"),
            "Day Volume":     day.get("v"),
            "Prev Close":     prev_day.get("c"),
            "Last Trade":     None,
            "Minute VWAP":    None,
            "Retrieved At":   run_ts,
        })

    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows)
    df = df.sort_values("Today Change %", ascending=False).head(TOP_N).reset_index(drop=True)
    df["Rank"] = range(1, len(df) + 1)
    print(f"  Found {len(df)} DM universe movers")
    return df


def main():
    # Part 1: Polygon top gainers (all stocks)
    print("=" * 65)
    print("  PART 1: POLYGON TOP GAINERS (all stocks)")
    print("=" * 65)
    df_polygon = fetch_top_gainers(limit=TOP_N)

    if not df_polygon.empty:
        print_summary(df_polygon)
        xlsx_path = write_dated_excel(df_polygon)
        append_run_log(df_polygon)
        print(f"Excel: {xlsx_path.resolve()}")
    else:
        print("No Polygon gainers passed filter.")

    # Part 2: DM Universe top movers
    print()
    print("=" * 65)
    print("  PART 2: DM UNIVERSE TOP MOVERS")
    print("=" * 65)
    df_dm = fetch_dm_universe_movers()

    if not df_dm.empty:
        print_summary(df_dm)
        # Save DM movers separately
        OUTPUT_DIR.mkdir(exist_ok=True)
        stamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        dm_path = OUTPUT_DIR / f"dm_universe_movers_{stamp}.csv"
        df_dm.to_csv(dm_path, index=False)
        print(f"DM Movers CSV: {dm_path.resolve()}")
    else:
        print("No DM universe mover data available.")


if __name__ == "__main__":
    main()
