import os
from datetime import datetime
from pathlib import Path
from dotenv import load_dotenv

load_dotenv()

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# Polygon/Massive market movers endpoint pattern.
# Set your key before running:
#   Windows PowerShell: $env:POLYGON_API_KEY="your_key_here"
#   macOS/Linux: export POLYGON_API_KEY="your_key_here"
API_KEY = os.getenv("POLYGON_API_KEY", "")
API_URL = "https://api.polygon.io/v2/snapshot/locale/us/markets/stocks/gainers"
OUTPUT_DIR = Path("market_snapshots_polygon")
TOP_N = 20  # Polygon's market movers endpoint returns the top 20.


def _pick_rows(payload):
    if isinstance(payload, list):
        return payload
    if isinstance(payload, dict):
        for key in ("tickers", "results"):
            if isinstance(payload.get(key), list):
                return payload[key]
    raise RuntimeError(f"Unexpected Polygon response shape: {payload}")


def fetch_top_gainers(limit: int = TOP_N) -> pd.DataFrame:
    if not API_KEY:
        raise RuntimeError(
            "Missing POLYGON_API_KEY environment variable. "
            "Set it first, then rerun the script."
        )

    resp = requests.get(API_URL, params={"apiKey": API_KEY}, timeout=30)
    resp.raise_for_status()
    payload = resp.json()
    rows = _pick_rows(payload)

    if not rows:
        raise RuntimeError("Polygon returned no gainers rows.")

    normalized = []
    run_ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for idx, item in enumerate(rows[:limit], start=1):
        day = item.get("day", {}) or {}
        prev_day = item.get("prevDay", {}) or {}
        last_trade = item.get("lastTrade", {}) or {}
        min_bar = item.get("min", {}) or {}

        symbol = item.get("ticker") or item.get("symbol")
        change_pct = item.get("todaysChangePerc")
        if change_pct is not None:
            change_pct = float(change_pct) / 100.0

        normalized.append(
            {
                "Rank": idx,
                "Symbol": symbol,
                "Price (Day Close)": day.get("c"),
                "Today Change": item.get("todaysChange"),
                "Today Change %": change_pct,
                "Day Open": day.get("o"),
                "Day High": day.get("h"),
                "Day Low": day.get("l"),
                "Day Volume": day.get("v"),
                "Prev Close": prev_day.get("c"),
                "Last Trade Price": last_trade.get("p"),
                "Minute VWAP": min_bar.get("vw"),
                "Minute Volume": min_bar.get("v"),
                "Retrieved At": run_ts,
                "Source": API_URL,
            }
        )

    return pd.DataFrame(normalized)


def autofit_worksheet(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 38)


def style_workbook(path: Path, row_count: int):
    wb = load_workbook(path)
    ws = wb["Top Gainers"]
    notes = wb["Notes"]

    ws["A1"] = "Top US Stock Gainers - Polygon"
    ws["A2"] = "Generated:"
    ws["B2"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["A3"] = "API Source:"
    ws["B3"] = API_URL

    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"].font = Font(bold=True)
    ws["A3"].font = Font(bold=True)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    header_row = 5
    last_row = header_row + row_count
    last_col = ws.max_column

    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    for row in ws.iter_rows(min_row=header_row + 1, max_row=last_row, min_col=1, max_col=last_col):
        for cell in row:
            cell.border = thin_border

    col_map = {cell.value: cell.column for cell in ws[header_row]}
    percent_cols = ["Today Change %"]
    money_cols = ["Price (Day Close)", "Today Change", "Day Open", "Day High", "Day Low", "Prev Close", "Last Trade Price", "Minute VWAP"]
    volume_cols = ["Day Volume", "Minute Volume"]

    for row_num in range(header_row + 1, last_row + 1):
        for name in percent_cols:
            if name in col_map:
                ws.cell(row_num, col_map[name]).number_format = "0.00%"
        for name in money_cols:
            if name in col_map:
                ws.cell(row_num, col_map[name]).number_format = "$0.00"
        for name in volume_cols:
            if name in col_map:
                ws.cell(row_num, col_map[name]).number_format = "#,##0"

    ws.freeze_panes = "A6"
    table_ref = f"A5:{get_column_letter(last_col)}{last_row}"
    table = Table(displayName="PolygonTopGainersTable", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    autofit_worksheet(ws)
    autofit_worksheet(notes)
    wb.save(path)


def write_dated_excel(df: pd.DataFrame) -> Path:
    OUTPUT_DIR.mkdir(exist_ok=True)
    stamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    output_path = OUTPUT_DIR / f"polygon_top_gainers_{stamp}.xlsx"

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Top Gainers", startrow=4, index=False)
        notes = pd.DataFrame(
            {
                "Notes": [
                    "Polygon API-backed market movers snapshot.",
                    "Each run creates a dated workbook instead of overwriting the prior file.",
                    "Set POLYGON_API_KEY before running.",
                ]
            }
        )
        notes.to_excel(writer, sheet_name="Notes", index=False)

    style_workbook(output_path, len(df))
    return output_path


def append_run_log(df: pd.DataFrame):
    OUTPUT_DIR.mkdir(exist_ok=True)
    log_path = OUTPUT_DIR / "polygon_top_gainers_history.csv"
    mode = "a" if log_path.exists() else "w"
    header = not log_path.exists()
    df.to_csv(log_path, mode=mode, header=header, index=False)


def main():
    df = fetch_top_gainers(limit=TOP_N)
    xlsx_path = write_dated_excel(df)
    append_run_log(df)
    print(f"Created: {xlsx_path.resolve()}")
    print(f"Updated history log: {(OUTPUT_DIR / 'polygon_top_gainers_history.csv').resolve()}")


if __name__ == "__main__":
    main()
