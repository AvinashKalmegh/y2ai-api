"""
top_gainers_dm_v3.py
===============================================================
DM UNIVERSE MOVERS — Daily Institutional Signal Report

APPROACH (v3 — complete rewrite):
  Previous versions pulled Polygon's top percentage gainers and tried
  to cross-reference against the DM universe. That approach failed
  because the Polygon gainers endpoint returns micro-cap penny stocks
  that are never in the DM universe.

  v3 flips the logic:
    1. Load the DM universe (DM_Latest.csv — ~594 institutional names)
    2. Fetch today's price snapshot for every ticker in that universe
    3. Find which DM universe names moved significantly today
    4. Rank by composite score (DM * price move)
    5. Output BACKED / WATCH / HEAD FAKE classification for each mover

  Result: every output row is guaranteed to be an institutional name
  in the DM universe. No penny stocks. No NO DATA rows.

SETUP:
  Windows PowerShell:
    $env:POLYGON_API_KEY="your_key_here"
    $env:DM_CSV_PATH="path/to/DM_Latest.csv"        # required
    $env:DM_HISTORY_PATH="path/to/DM_History.csv"   # optional, enables lookback

  macOS/Linux:
    export POLYGON_API_KEY="your_key_here"
    export DM_CSV_PATH="path/to/DM_Latest.csv"
    export DM_HISTORY_PATH="path/to/DM_History.csv"

  Run:
    python top_gainers_dm_v3.py

OUTPUT:
  market_snapshots_polygon/
    dm_universe_movers_YYYY-MM-DD_HH-MM-SS.xlsx  (dated snapshot)
    dm_universe_movers_history.csv               (accumulates daily)

HISTORY FILE:
  Starts clean from April 8, 2026.
  Each run appends the results to dm_universe_movers_history.csv.
  Over time this builds a searchable record of which DM universe names
  moved each day and what their signal was.
===============================================================
"""

import os
import math
import time
from datetime import datetime, timedelta
from pathlib import Path

from dotenv import load_dotenv
load_dotenv()

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ============================================================
# CONFIGURATION
# ============================================================

API_KEY       = os.getenv("POLYGON_API_KEY", "")
OUTPUT_DIR    = Path("market_snapshots_polygon")

SUPABASE_URL     = os.getenv("SUPABASE_URL", "")
SUPABASE_KEY     = os.getenv("SUPABASE_KEY", "")

# -- Mover thresholds ------------------------------------------
# A name must move at least this much today to appear in output.
# Keeps the list focused on names with real price action.
MIN_PRICE_MOVE_PCT = float(os.getenv("DM_MIN_MOVE",    "2.0"))   # % price move today
MIN_VOLUME         = float(os.getenv("DM_MIN_VOLUME",  "500000"))# minimum day volume
MIN_PRICE          = float(os.getenv("DM_MIN_PRICE",   "5.0"))   # minimum price ($)

# -- Output control --------------------------------------------
TOP_N         = int(os.getenv("DM_TOP_N", "25"))     # max names in output
LOOKBACK_DAYS = int(os.getenv("DM_LOOKBACK_DAYS", "90"))

# -- DM signal thresholds -------------------------------------
DM_STRONG   = 65   # BACKED — institutional accumulation confirmed
DM_NEUTRAL  = 40   # WATCH  — mixed signal
# < DM_NEUTRAL = HEAD FAKE — price up but institutions leaving

# -- Polygon API -----------------------------------------------
# Snapshot endpoint accepts up to 200 tickers per call
POLYGON_SNAPSHOT_URL = (
    "https://api.polygon.io/v2/snapshot/locale/us/markets/stocks/tickers"
)
POLYGON_BATCH_SIZE   = 150   # conservative — stay well under 200
POLYGON_TIMEOUT      = 30


# ============================================================
# SUPABASE CLIENT
# ============================================================

def _get_supabase():
    if not SUPABASE_URL or not SUPABASE_KEY:
        return None
    from supabase import create_client
    return create_client(SUPABASE_URL, SUPABASE_KEY)


# ============================================================
# DM DATA LOADERS (from Supabase)
# ============================================================

def load_dm_latest() -> pd.DataFrame:
    """
    Load DM universe from Supabase dm_latest table.
    Returns a DataFrame with columns: Ticker, DM, DM_Change, Phase, Close
    """
    client = _get_supabase()
    if not client:
        raise RuntimeError(
            "SUPABASE_URL/SUPABASE_KEY not set. Cannot load DM universe."
        )

    rows = []
    offset, page = 0, 5000
    while True:
        r = (client.table('dm_latest')
             .select('ticker,dm_smoothed,dm_change,phase,close')
             .range(offset, offset + page - 1)
             .execute())
        rows.extend(r.data)
        if len(r.data) < page:
            break
        offset += page

    if not rows:
        raise RuntimeError("Supabase dm_latest returned no rows.")

    df = pd.DataFrame(rows)
    df.rename(columns={
        'ticker': 'Ticker', 'dm_smoothed': 'DM',
        'dm_change': 'DM_Change', 'phase': 'Phase', 'close': 'Close'
    }, inplace=True)
    df["Ticker"]    = df["Ticker"].str.strip().str.upper()
    df["DM"]        = pd.to_numeric(df["DM"], errors="coerce")
    df["DM_Change"] = pd.to_numeric(df["DM_Change"], errors="coerce")
    df["Close"]     = pd.to_numeric(df["Close"], errors="coerce")
    df = df.dropna(subset=["Ticker", "DM"])

    print(f"[DM Universe] Loaded {len(df)} tickers from Supabase dm_latest")
    return df


def load_dm_history() -> pd.DataFrame | None:
    """
    Load DM history from Supabase dm_history table for lookback analysis.
    Returns None if unavailable.
    """
    client = _get_supabase()
    if not client:
        print("[DM History] Supabase not available -- lookback disabled.")
        return None

    try:
        cutoff = (datetime.now() - timedelta(days=int(LOOKBACK_DAYS * 1.5))).strftime('%Y-%m-%d')
        print(f"[DM History] Loading from Supabase (since {cutoff})...")

        rows = []
        offset, page = 0, 10000
        while True:
            r = (client.table('dm_history')
                 .select('date,ticker,dm_smoothed,phase,close')
                 .gte('date', cutoff)
                 .order('date')
                 .range(offset, offset + page - 1)
                 .execute())
            rows.extend(r.data)
            if len(r.data) < page:
                break
            offset += page

        if not rows:
            print("[DM History] No rows returned -- lookback disabled.")
            return None

        df = pd.DataFrame(rows)
        df.rename(columns={
            'dm_smoothed': 'DM', 'ticker': 'Ticker',
            'date': 'Date', 'close': 'Close', 'phase': 'Phase'
        }, inplace=True)
        df['Date']   = pd.to_datetime(df['Date'])
        df['DM']     = pd.to_numeric(df['DM'], errors='coerce')
        df['Close']  = pd.to_numeric(df['Close'], errors='coerce')
        df['Ticker'] = df['Ticker'].str.strip().str.upper()
        df = df.sort_values(['Ticker', 'Date'])

        print(f"[DM History] Loaded {len(df):,} rows -- "
              f"{df['Ticker'].nunique()} tickers -- "
              f"{df['Date'].min().date()} -> {df['Date'].max().date()}")
        return df
    except Exception as e:
        print(f"[WARN] Failed to load DM history: {e}")
        return None


# ============================================================
# POLYGON PRICE FETCH — batched snapshot of DM universe
# ============================================================

def fetch_universe_snapshots(tickers: list[str]) -> dict[str, dict]:
    """
    Fetch today's price snapshot for a list of tickers from Polygon.
    Returns dict: ticker -> {price, change_pct, volume, open, high, low, vwap}
    Processes in batches of POLYGON_BATCH_SIZE to stay within API limits.
    """
    if not API_KEY:
        raise RuntimeError(
            "Missing POLYGON_API_KEY environment variable.\n"
            "Set it first, then rerun the script."
        )

    result    = {}
    n_batches = math.ceil(len(tickers) / POLYGON_BATCH_SIZE)

    print(f"[Polygon] Fetching snapshots for {len(tickers)} universe tickers "
          f"in {n_batches} batch(es)...")

    for batch_idx in range(n_batches):
        batch  = tickers[batch_idx * POLYGON_BATCH_SIZE :
                         (batch_idx + 1) * POLYGON_BATCH_SIZE]
        joined = ",".join(batch)

        try:
            resp = requests.get(
                POLYGON_SNAPSHOT_URL,
                params={"tickers": joined, "apiKey": API_KEY},
                timeout=POLYGON_TIMEOUT,
            )
            resp.raise_for_status()
            data  = resp.json()
            snaps = data.get("tickers") or data.get("results") or []

            for snap in snaps:
                tk  = snap.get("ticker", "")
                day = snap.get("day")     or {}
                mn  = snap.get("min")     or {}

                price      = day.get("c") or snap.get("lastTrade", {}).get("p")
                volume     = day.get("v")
                open_px    = day.get("o")
                high_px    = day.get("h")
                low_px     = day.get("l")
                vwap       = mn.get("vw")
                prev_close = (snap.get("prevDay") or {}).get("c")
                change_pct = snap.get("todaysChangePerc")
                change_amt = snap.get("todaysChange")

                if price is None:
                    continue

                result[tk.upper()] = {
                    "price":      float(price),
                    "change_pct": float(change_pct) / 100.0 if change_pct is not None else None,
                    "change_amt": float(change_amt)          if change_amt is not None else None,
                    "volume":     float(volume)              if volume     is not None else 0.0,
                    "open":       float(open_px)             if open_px    is not None else None,
                    "high":       float(high_px)             if high_px    is not None else None,
                    "low":        float(low_px)              if low_px     is not None else None,
                    "vwap":       float(vwap)                if vwap       is not None else None,
                    "prev_close": float(prev_close)          if prev_close is not None else None,
                }

        except Exception as e:
            print(f"[Polygon] Batch {batch_idx + 1} error: {e}")

        # Brief pause between batches to respect rate limits
        if batch_idx < n_batches - 1:
            time.sleep(0.25)

    fetched = len(result)
    print(f"[Polygon] Received prices for {fetched} of {len(tickers)} tickers "
          f"({fetched/len(tickers)*100:.0f}%)")
    return result


# ============================================================
# LOOKBACK ANALYSIS
# ============================================================

def classify_dm(dm_score: float | None) -> str:
    if dm_score is None or pd.isna(dm_score):
        return "NO DATA"
    if dm_score >= DM_STRONG:
        return "BACKED"
    if dm_score >= DM_NEUTRAL:
        return "WATCH"
    return "HEAD FAKE"


def find_entry_signal(history_df: pd.DataFrame, ticker: str,
                      current_price: float) -> dict:
    """
    Scan the lookback window for the first date DM crossed DM_STRONG.
    Returns entry analysis including simulated return.
    """
    empty = {
        "entry_date": None, "entry_dm": None, "entry_price": None,
        "days_held": None, "simulated_return": None,
        "signal_type": "NO_HISTORY", "peak_dm": None, "days_above_gate": None,
    }

    if history_df is None:
        return empty

    tk_df = history_df[history_df["Ticker"] == ticker].copy()
    if tk_df.empty:
        return {**empty, "signal_type": "NOT_IN_HISTORY"}

    cutoff = datetime.now() - timedelta(days=LOOKBACK_DAYS * 1.5)
    tk_df  = tk_df[tk_df["Date"] >= cutoff].reset_index(drop=True)
    if tk_df.empty:
        return {**empty, "signal_type": "NO_HISTORY"}

    above      = tk_df[tk_df["DM"] >= DM_STRONG]
    peak_dm    = float(tk_df["DM"].max())
    days_above = int((tk_df["DM"] >= DM_STRONG).sum())

    if above.empty:
        return {**empty, "signal_type": "NEVER_CROSSED",
                "peak_dm": peak_dm, "days_above_gate": 0}

    entry_row   = above.iloc[0]
    entry_date  = entry_row["Date"]
    entry_dm    = float(entry_row["DM"])
    entry_price = (float(entry_row["Close"])
                   if "Close" in entry_row and pd.notna(entry_row["Close"])
                   else None)
    days_held   = (datetime.now() - entry_date).days
    sim_return  = None
    if entry_price and entry_price > 0 and current_price:
        sim_return = (current_price - entry_price) / entry_price * 100

    return {
        "entry_date":       entry_date.strftime("%Y-%m-%d"),
        "entry_dm":         round(entry_dm, 1),
        "entry_price":      round(entry_price, 2) if entry_price else None,
        "days_held":        days_held,
        "simulated_return": round(sim_return, 1) if sim_return is not None else None,
        "signal_type":      "STRONG_ENTRY",
        "peak_dm":          round(peak_dm, 1),
        "days_above_gate":  days_above,
    }


# ============================================================
# CORE: BUILD MOVERS TABLE
# ============================================================

def build_movers(dm_df: pd.DataFrame,
                 snapshots: dict[str, dict],
                 history_df: pd.DataFrame | None) -> pd.DataFrame:
    """
    Cross-reference DM universe with today's price snapshots.
    Returns a DataFrame of tickers that moved significantly today,
    ranked by composite score (DM × |price_move|).
    """
    run_ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    rows   = []

    for _, dm_row in dm_df.iterrows():
        ticker = dm_row["Ticker"]
        snap   = snapshots.get(ticker)

        if snap is None:
            continue  # No price data today — skip

        price      = snap["price"]
        change_pct = snap["change_pct"]
        volume     = snap["volume"]

        # -- Apply mover filters ------------------------------
        if price   < MIN_PRICE:
            continue
        if volume  < MIN_VOLUME:
            continue
        if change_pct is None:
            continue
        if abs(change_pct) * 100 < MIN_PRICE_MOVE_PCT:
            continue   # Not moving enough today

        dm_score  = dm_row.get("DM")
        dm_change = dm_row.get("DM_Change") if "DM_Change" in dm_row else None
        phase     = dm_row.get("Phase", "")  if "Phase"    in dm_row else ""
        signal    = classify_dm(dm_score)

        # -- Lookback -----------------------------------------
        entry = find_entry_signal(history_df, ticker, price)

        # -- Composite rank score -----------------------------
        # Higher DM + bigger move = higher priority
        # Use abs(change_pct) so declining names (potential shorts) also surface
        dm_for_score = dm_score if dm_score is not None else 0
        composite    = dm_for_score * abs(change_pct) * 100

        rows.append({
            "Composite":         composite,           # for sorting only
            "Rank":              0,                   # assigned after sort
            "Symbol":            ticker,
            "Signal":            signal,
            "DM Score":          dm_score,
            "DM Change":         dm_change,
            "DM Phase":          phase,
            "Price":             price,
            "Today Change %":    change_pct,
            "Today Change $":    snap["change_amt"],
            "Day Open":          snap["open"],
            "Day High":          snap["high"],
            "Day Low":           snap["low"],
            "Day Volume":        snap["volume"],
            "Prev Close":        snap["prev_close"],
            "Minute VWAP":       snap["vwap"],
            # Lookback
            "LB Signal":         entry["signal_type"],
            "LB Entry Date":     entry["entry_date"],
            "LB Entry DM":       entry["entry_dm"],
            "LB Entry Price":    entry["entry_price"],
            "LB Days Held":      entry["days_held"],
            "LB Sim Return %":   entry["simulated_return"],
            "LB Peak DM":        entry["peak_dm"],
            "LB Days Above 65":  entry["days_above_gate"],
            "Retrieved At":      run_ts,
        })

    if not rows:
        print("[WARN] No DM universe names matched today's mover criteria.")
        return pd.DataFrame()

    df = (pd.DataFrame(rows)
            .sort_values("Composite", ascending=False)
            .head(TOP_N)
            .reset_index(drop=True))

    df["Rank"] = range(1, len(df) + 1)
    df = df.drop(columns=["Composite"])   # internal — not in output

    print(f"[Movers] {len(df)} DM universe names moved >={MIN_PRICE_MOVE_PCT}% "
          f"today with volume >={int(MIN_VOLUME):,}")
    return df


# ============================================================
# CONSOLE SUMMARY
# ============================================================

def print_summary(df: pd.DataFrame):
    if df.empty:
        print("\n  No movers found today matching criteria.")
        return

    print()
    print("=" * 80)
    print(f"  DM UNIVERSE MOVERS  {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print(f"  Move gate: >={MIN_PRICE_MOVE_PCT}% today | "
          f"Volume >={int(MIN_VOLUME):,} | Price >=${MIN_PRICE:.0f}")
    print(f"  Signal gate: BACKED = DM >={DM_STRONG} | "
          f"WATCH = {DM_NEUTRAL}-{DM_STRONG} | HEAD FAKE = <{DM_NEUTRAL}")
    print("=" * 80)

    for signal_label, group_label in [
        ("BACKED",    f"OK  BACKED — DM >={DM_STRONG} (institutional flow confirmed)"),
        ("WATCH",     f"WATCH  WATCH  — DM {DM_NEUTRAL}-{DM_STRONG} (monitor, don't chase)"),
        ("HEAD FAKE", f"FAKE  HEAD FAKE — DM <{DM_NEUTRAL} (price up, institutions leaving)"),
    ]:
        grp = df[df["Signal"] == signal_label]
        if grp.empty:
            continue
        print(f"\n  {group_label}  ({len(grp)} names):")
        hdr = (f"    {'Sym':<7} {'Chg%':>7}  {'DM':>6}  {'DM Chg':>8}  {'Phase':<10}  "
               f"{'LB Signal':<16} {'Entry Dt':<12} {'SimRet%':>8}")
        print(hdr)
        print("    " + "-" * (len(hdr) - 4))
        for _, r in grp.iterrows():
            chg    = f"{r['Today Change %']:+.1%}"   if pd.notna(r.get("Today Change %")) else " N/A"
            dm     = f"{r['DM Score']:.1f}"           if pd.notna(r.get("DM Score"))       else "  --"
            dmc    = f"{r['DM Change']:+.1f}"         if pd.notna(r.get("DM Change"))      else "    --"
            phase  = str(r.get("DM Phase") or "")[:10]
            lb_sig = str(r.get("LB Signal") or "--")[:16]
            edate  = str(r.get("LB Entry Date") or "--")
            simret = (f"{r['LB Sim Return %']:+.1f}%"
                      if pd.notna(r.get("LB Sim Return %")) else "     --")
            print(f"    {r['Symbol']:<7} {chg:>7}  {dm:>6}  {dmc:>8}  "
                  f"{phase:<10}  {lb_sig:<16} {edate:<12} {simret:>8}")

    # Summary stats
    backed = df[df["Signal"] == "BACKED"]
    watch  = df[df["Signal"] == "WATCH"]
    hf     = df[df["Signal"] == "HEAD FAKE"]
    lb_pos = df[df["LB Sim Return %"] > 0]
    lb_all = df[df["LB Sim Return %"].notna()]

    print()
    print(f"  SIGNAL BREAKDOWN: {len(backed)} backed | "
          f"{len(watch)} watch | {len(hf)} head fakes | "
          f"{len(df) - len(backed) - len(watch) - len(hf)} other")
    if not lb_all.empty:
        avg = lb_all["LB Sim Return %"].mean()
        print(f"  LOOKBACK ({LOOKBACK_DAYS}d): avg sim return {avg:+.1f}% | "
              f"{len(lb_pos)}/{len(lb_all)} positive")
    print("=" * 80)
    print()


# ============================================================
# EXCEL OUTPUT
# ============================================================

SIGNAL_COLORS = {
    "BACKED":    "C6EFCE",
    "WATCH":     "FFEB9C",
    "HEAD FAKE": "FFC7CE",
    "NO DATA":   "D9D9D9",
}


def autofit_worksheet(ws):
    for col_cells in ws.columns:
        max_len    = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            val     = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 40)


def style_workbook(path: Path, row_count: int):
    wb    = load_workbook(path)
    ws    = wb["DM Movers"]
    notes = wb["Notes"]

    ws["A1"] = "DM Universe Movers — Institutional Signal Report"
    ws["A2"] = "Generated:"
    ws["B2"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["A3"] = "Move gate:"
    ws["B3"] = (f">={MIN_PRICE_MOVE_PCT}% daily move | "
                f"Volume >={int(MIN_VOLUME):,} | Price >=${MIN_PRICE:.0f}")
    ws["A4"] = "Signal gate:"
    ws["B4"] = (f"BACKED = DM >={DM_STRONG} | "
                f"WATCH = {DM_NEUTRAL}-{DM_STRONG} | "
                f"HEAD FAKE = <{DM_NEUTRAL}")

    ws["A1"].font = Font(bold=True, size=14)
    for r in ["A2","A3","A4"]:
        ws[r].font = Font(bold=True)

    hdr_fill    = PatternFill("solid", fgColor="1F4E79")
    lb_hdr_fill = PatternFill("solid", fgColor="2E4057")
    hdr_font    = Font(color="FFFFFF", bold=True)
    thin        = Border(
        left=Side(style="thin", color="D9D9D9"),  right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),   bottom=Side(style="thin", color="D9D9D9"),
    )

    HEADER_ROW = 6
    last_row   = HEADER_ROW + row_count
    last_col   = ws.max_column

    col_map = {cell.value: cell.column for cell in ws[HEADER_ROW]}

    for cell in ws[HEADER_ROW]:
        is_lb     = str(cell.value or "").startswith("LB ")
        cell.fill = lb_hdr_fill if is_lb else hdr_fill
        cell.font = hdr_font
        cell.border = thin

    signal_col = col_map.get("Signal")
    lb_ret_col = col_map.get("LB Sim Return %")
    pct_cols   = {"Today Change %"}
    money_cols = {"Price","Today Change $","Day Open","Day High","Day Low",
                  "Prev Close","Minute VWAP","LB Entry Price"}
    vol_cols   = {"Day Volume"}

    for row_num in range(HEADER_ROW + 1, last_row + 1):
        for col_num in range(1, last_col + 1):
            ws.cell(row_num, col_num).border = thin

        if signal_col:
            sv = ws.cell(row_num, signal_col).value or ""
            fc = SIGNAL_COLORS.get(sv)
            if fc:
                fill = PatternFill("solid", fgColor=fc)
                for c in range(1, last_col + 1):
                    ws.cell(row_num, c).fill = fill

        if lb_ret_col:
            v = ws.cell(row_num, lb_ret_col).value
            if v is not None:
                try:
                    vf = float(v)
                    ws.cell(row_num, lb_ret_col).fill = PatternFill(
                        "solid", fgColor="C6EFCE" if vf >= 0 else "FFC7CE")
                    ws.cell(row_num, lb_ret_col).number_format = "+0.0%;-0.0%"
                except (ValueError, TypeError):
                    pass

        for name in pct_cols:
            if name in col_map:
                ws.cell(row_num, col_map[name]).number_format = "+0.00%;-0.00%"
        for name in money_cols:
            if name in col_map:
                ws.cell(row_num, col_map[name]).number_format = "$#,##0.00"
        for name in vol_cols:
            if name in col_map:
                ws.cell(row_num, col_map[name]).number_format = "#,##0"

    ws.freeze_panes = f"A{HEADER_ROW + 1}"
    table_ref = f"A{HEADER_ROW}:{get_column_letter(last_col)}{last_row}"
    table     = Table(displayName="DMMoversTable", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(table)
    autofit_worksheet(ws)
    autofit_worksheet(notes)
    wb.save(path)


def write_dated_excel(df: pd.DataFrame) -> Path:
    OUTPUT_DIR.mkdir(exist_ok=True)
    stamp       = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    output_path = OUTPUT_DIR / f"dm_universe_movers_{stamp}.xlsx"

    notes_rows = [
        "DM Universe Movers — starts from DM universe, not Polygon gainers list.",
        "Every output row is guaranteed to be an institutional DM universe name.",
        "",
        f"Move gate: >={MIN_PRICE_MOVE_PCT}% price change today",
        f"Volume gate: >={int(MIN_VOLUME):,} daily volume",
        f"Price gate: >=${MIN_PRICE:.0f}",
        f"Max output rows: {TOP_N} (ranked by DM × |price move|)",
        "",
        "SIGNAL DEFINITIONS:",
        f"  BACKED    = DM >={DM_STRONG}: institutional accumulation confirmed. Real move.",
        f"  WATCH     = DM {DM_NEUTRAL}-{DM_STRONG}: mixed signal. Monitor, do not chase.",
        f"  HEAD FAKE = DM <{DM_NEUTRAL}: price up but institutions are leaving. Avoid.",
        "",
        "LOOKBACK COLUMNS (LB prefix):",
        f"  Lookback window: {LOOKBACK_DAYS} trading days",
        "  LB Signal:",
        "    STRONG_ENTRY   = DM crossed 65 at some point in the lookback window.",
        "    NEVER_CROSSED  = DM never reached 65 in the lookback window.",
        "    NOT_IN_HISTORY = ticker not in DM history file.",
        "    NO_HISTORY     = history file not configured.",
        "  LB Entry Date:   first date DM >= 65 in the window.",
        "  LB Entry Price:  closing price on that date.",
        "  LB Sim Return %: (today price - entry price) / entry price.",
        "  LB Days Held:    calendar days from entry to today.",
        "  LB Peak DM:      highest DM in the lookback window.",
        "  LB Days Above 65: sessions where DM >= 65 in window.",
        "",
        "DATA SOURCES:",
        "  DM Universe: Supabase dm_latest table",
        "  DM History:  Supabase dm_history table",
        "  Prices: Polygon.io snapshot API",
        "",
        "HISTORY FILE: dm_universe_movers_history.csv",
        "  Clean start: April 8, 2026.",
        "  Accumulates one run per day. Grows over time into signal validation dataset.",
    ]

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Leave 5 rows at top for metadata (written by style_workbook)
        df.to_excel(writer, sheet_name="DM Movers", startrow=5, index=False)
        pd.DataFrame({"Notes": notes_rows}).to_excel(
            writer, sheet_name="Notes", index=False)

    style_workbook(output_path, len(df))
    return output_path


# ============================================================
# HISTORY LOG
# ============================================================

def append_run_log(df: pd.DataFrame):
    """Append today's movers to the running history CSV."""
    if df.empty:
        return
    OUTPUT_DIR.mkdir(exist_ok=True)
    log_path = OUTPUT_DIR / "dm_universe_movers_history.csv"
    mode     = "a" if log_path.exists() else "w"
    header   = not log_path.exists()
    df.to_csv(log_path, mode=mode, header=header, index=False)
    print(f"[History] Appended {len(df)} rows -> {log_path.resolve()}")


# ============================================================
# MAIN
# ============================================================

def main():
    print()
    print(">>  DM Universe Movers v3")
    print(f"   {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()

    # 1. Load DM universe from Supabase
    dm_df      = load_dm_latest()
    history_df = load_dm_history()

    # 2. Fetch price snapshots for entire universe
    tickers   = dm_df["Ticker"].tolist()
    snapshots = fetch_universe_snapshots(tickers)

    # 3. Build movers table
    df = build_movers(dm_df, snapshots, history_df)

    # 4. Print summary
    print_summary(df)

    if df.empty:
        print("No output files written — no movers found today.")
        return

    # 5. Write Excel
    xlsx_path = write_dated_excel(df)
    print(f"Excel:   {xlsx_path.resolve()}")

    # 6. Append to history
    append_run_log(df)
    log_path = OUTPUT_DIR / "dm_universe_movers_history.csv"
    print(f"History: {log_path.resolve()}")
    print()
    print("Done.")


if __name__ == "__main__":
    main()
